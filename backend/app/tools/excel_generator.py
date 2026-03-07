"""Excel generation builtin tool.

Input params:
{
  "filename": "报表",
  "sheets": [
    {
      "name": "Sheet1",
      "headers": ["列1", "列2", "列3"],
      "rows": [["A", "B", "C"], ["D", "E", "F"]]
    }
  ]
}

Output: {"file_id": "xxx.xlsx", "download_url": "/api/files/xxx.xlsx"}
"""
from __future__ import annotations

import os
import uuid
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

_UPLOAD_DIR = os.environ.get("UPLOAD_DIR", "./uploads")
_GENERATED_DIR = Path(_UPLOAD_DIR) / "generated"


def execute(params: dict) -> dict:
    """Generate an XLSX file and return file_id + download_url."""
    _GENERATED_DIR.mkdir(parents=True, exist_ok=True)

    filename_base = params.get("filename", "报表")
    sheets_data = params.get("sheets", [])

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    if not sheets_data:
        # Create an empty default sheet
        sheets_data = [{"name": "Sheet1", "headers": [], "rows": []}]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for sheet_data in sheets_data:
        ws = wb.create_sheet(title=sheet_data.get("name", "Sheet"))
        headers = sheet_data.get("headers", [])
        rows = sheet_data.get("rows", [])

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit column widths (approximate)
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    file_id = f"{uuid.uuid4().hex}.xlsx"
    file_path = _GENERATED_DIR / file_id
    wb.save(str(file_path))

    total_rows = sum(len(s.get("rows", [])) for s in sheets_data)
    return {
        "file_id": file_id,
        "download_url": f"/api/files/{file_id}",
        "filename": f"{filename_base}.xlsx",
        "message": f"Excel已生成，共{len(sheets_data)}个Sheet，{total_rows}行数据",
    }
